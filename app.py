import os
import pandas as pd
from tkinter import *
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# -------------- Main Excel Image Function ----------------
def generate_excel_with_images(input_excel, images_base_folder, output_excel):
    if not os.path.exists(input_excel):
        raise FileNotFoundError("Excel file not found.")

    df = pd.read_excel(input_excel)
    df.to_excel(output_excel, index=False)

    wb = load_workbook(output_excel)
    ws = wb.active
    image_col = ws.max_column + 1
    ws.cell(row=1, column=image_col, value="Image")

    for i, row in df.iterrows():
        folder_name = str(row['Folder']).strip()
        folder_path = os.path.join(images_base_folder, folder_name)

        if not os.path.isdir(folder_path):
            continue

        image_files = sorted([
            f for f in os.listdir(folder_path)
            if f.lower().endswith(('.jpg', '.jpeg', '.png'))
        ])

        row_index = i % 5
        if row_index < len(image_files):
            img_path = os.path.join(folder_path, image_files[row_index])
            try:
                img = ExcelImage(img_path)
                cell = f"{chr(64 + image_col)}{i + 2}"
                ws.add_image(img, cell)
            except Exception as e:
                print("Image error:", e)

    wb.save(output_excel)

# ---------------- UI Functions ----------------
def browse_file(entry): entry.delete(0, END); entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))
def browse_folder(entry): entry.delete(0, END); entry.insert(0, filedialog.askdirectory())
def browse_output(entry): entry.delete(0, END); entry.insert(0, filedialog.asksaveasfilename(defaultextension=".xlsx"))
def on_drop(event, entry): entry.delete(0, END); entry.insert(0, event.data.strip('{}'))

def run_app():
    if not input_entry.get() or not folder_entry.get() or not output_entry.get():
        messagebox.showerror("Missing Input", "Please fill in all fields.")
        return
    try:
        generate_excel_with_images(input_entry.get(), folder_entry.get(), output_entry.get())
        messagebox.showinfo("✅ Success", f"Excel saved:\n{output_entry.get()}")
    except Exception as e:
        messagebox.showerror("❌ Error", str(e))

# ---------------- Main UI ----------------
root = TkinterDnD.Tk()
root.title("🦉 Excel Image Inserter")
root.geometry("640x400")
root.configure(bg="#2a2d35")
root.resizable(False, False)

style_font = ("Segoe UI", 11)
title_font = ("Segoe UI", 14, "bold")

def make_row(parent, label_text, row, entry_ref, browse_fn):
    Label(parent, text=label_text, font=style_font, bg="#2a2d35", fg="white").grid(row=row, column=0, sticky="w", pady=10, padx=10)
    entry = Entry(parent, width=55, bg="#3b3e48", fg="white", relief=FLAT, font=style_font, insertbackground="white")
    entry.grid(row=row, column=1, padx=(0, 10))
    entry.drop_target_register(DND_FILES)
    entry.dnd_bind("<<Drop>>", lambda e: on_drop(e, entry))
    browse_btn = Button(parent, text="Browse", command=lambda: browse_fn(entry),
                        bg="#cdd0ea", fg="#1e1e1e", relief=FLAT, font=("Segoe UI", 10, "bold"), padx=12, pady=4)
    browse_btn.grid(row=row, column=2)
    entry_ref.append(entry)

# Header
Label(root, text="🦉 Excel Image Inserter", font=title_font, bg="#2a2d35", fg="#cdd0ea").pack(pady=(20, 10))

frame = Frame(root, bg="#2a2d35")
frame.pack()

input_entry = []
folder_entry = []
output_entry = []

make_row(frame, "Input Excel File:", 0, input_entry, browse_file)
make_row(frame, "Images Base Folder:", 1, folder_entry, browse_folder)
make_row(frame, "Output Excel File:", 2, output_entry, browse_output)

input_entry = input_entry[0]
folder_entry = folder_entry[0]
output_entry = output_entry[0]

# Run Button
Button(root, text="🚀 Generate Excel", command=run_app,
       bg="#57cc99", fg="black", font=("Segoe UI", 11, "bold"),
       relief=FLAT, padx=20, pady=10, cursor="hand2").pack(pady=30)

root.mainloop()
